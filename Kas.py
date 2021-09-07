

"""pip install Pillow for Python 3.7 imagery"""
 
from tkinter import *
from tkinter import ttk
import tkinter
from tkinter import ttk
from PIL import ImageTk, Image
import os
import numpy as np
from os import path
import shutil
from pathlib import Path
import pandas as pd
import xlrd
import xlwt
from xlwt import Workbook
from tkinter import messagebox
from tkinter import Label

import openpyxl
from openpyxl import load_workbook
import sys



import xlsxwriter


 
main = Tk()
"""ErrorRun"""
try:
    wbTest = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\Teller.xlsx')
except:
    messagebox.showinfo("No Go","Teller open")
    sys.exit(1)

try:
    wbTest1 = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\... Taken.xlsm')
except:
    messagebox.showinfo("No Go","MelanieTaken open")
    sys.exit(1)

try:
    fo = open("KasbladDemo.xls", "wb")
    fo.close()
except:
    messagebox.showinfo("No Go","KasbladDemo open")
    sys.exit(1)
    
    
    


"""Databases"""
MainDump = []


KlantDatabase = []
ExpertMDB = []
KlantDB = []
wb66 = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\... Taken.xlsm', read_only=True)
sheet66 = wb66['Database']
wb99 = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\Teller.xlsx', read_only=True)
sheet99 = wb99['Data']

EDB = 'A'
KlantDBEnu = 'B'
mmm = 0
for i in range(2,1100):
    try:
        EDB = EDB +str(mmm)
        if sheet66[EDB].value == "":
            pass
        else:
            ExpertMDB.append(str(sheet66[EDB].value))
        EDB = 'A'
        mmm = mmm + 1
    except:
        mmm = mmm + 1

mm = 0
for i in range(2,1100):
    try:
        KlantDBEnu = KlantDBEnu + str(mm)
        if sheet66[KlantDBEnu].value == "":
            pass
        else:
            KlantDB.append(str(sheet66[KlantDBEnu].value))
        KlantDBEnu = 'B'
        mm = mm + 1
    except:
        mm = mm + 1

wb66.close()

"""GUI"""

main.title('KAS')
main.geometry('1000x600')

"""define"""

"""functie Maak rapport"""
        
def CreateExcel():
    wb99 = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\Teller.xlsx', read_only=True)
    sheet99 = wb99['Data']

    KN = []
    KNenu = 'B'
    k = 2
    for i in range(2,5):
        try:
            KNenu = KNenu+str(k)
            KN.append(str(sheet99[KNenu].value))
            KNenu = 'B'
            k = k + 1
        except:
            pass
        
    KlantNrNaam = []
    wb2 = load_workbook(filename = r'C:\Users\...s\OneDrive - Gamesgroup\Documenten\Kasblad.xlsx')
    sheet2 = wb2['Sheet1']
    KlantNNEnu = 'R'
    for i in range(2, 100):
        KlantNNEnu = KlantNNEnu + str(i)
        i = i + 1
        Value = sheet2[KlantNNEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                KlantNrNaam.append(Value)
                MainDump.append(Value)
        KlantNNEnu = 'R'

    Taks = []
    wb = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\Kasblad.xlsx')
    sheet = wb['Sheet1']
    TaksEnu = 'H'
    for i in range(2, 100):
        TaksEnu = TaksEnu + str(i)
        i = i + 1
        Value = sheet[TaksEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                Taks.append(Value)
                MainDump.append(Value)
        TaksEnu = 'H'

    Schuldtaks = []
    wb1 = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\Kasblad.xlsx')
    sheet1 = wb1['Sheet1']

    SchuldTaksEnu = 'M'
    for i in range(2, 100):
        SchuldTaksEnu = SchuldTaksEnu + str(i)
        i = i + 1
        Value = sheet1[SchuldTaksEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                Schuldtaks.append(Value)
                MainDump.append(Value)
        SchuldTaksEnu = 'M'
    
    Provisie = []
    ProvisieEnu = 'J'
    for i in range(2, 100):
        ProvisieEnu = ProvisieEnu + str(i)
        i = i + 1
        Value = sheet1[ProvisieEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                Provisie.append(Value)
                MainDump.append(Value)
        ProvisieEnu = 'J'

    AfbetalingTaks = []
    ATaksEnu = 'K'
    for i in range(2, 100):
        ATaksEnu = ATaksEnu + str(i)
        i = i + 1
        Value = sheet1[ATaksEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                AfbetalingTaks.append(Value)
                MainDump.append(Value)
        ATaksEnu = 'K'

    SchuldLichting = []
    SLichtingEnu = 'L'
    for i in range(2, 100):
        SLichtingEnu = SLichtingEnu + str(i)
        i = i + 1
        Value = sheet1[SLichtingEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                SchuldLichting.append(Value)
                MainDump.append(Value)
        SLichtingEnu = 'L'

    AfbetalingLichting = []
    ALichtingEnu = 'I'
    for i in range(2, 100):
        ALichtingEnu = ALichtingEnu + str(i)
        i = i + 1
        Value = sheet1[ALichtingEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                AfbetalingLichting.append(Value)
                MainDump.append(Value)
        ALichtingEnu = 'I'

    Revenue = []
    RevenueEnu = 'F'
    for i in range(2, 100):
        RevenueEnu = RevenueEnu + str(i)
        i = i + 1
        Value = sheet1[RevenueEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                Revenue.append(Value)
                MainDump.append(Value)
        RevenueEnu = 'F'

    Onkosten = []
    OnkostenEnu = 'G'
    for i in range(2, 100):
        OnkostenEnu = OnkostenEnu + str(i)
        i = i + 1
        Value = sheet1[OnkostenEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                Onkosten.append(Value)
                MainDump.append(Value)
        OnkostenEnu = 'G'

    Bancontact = []
    BancontactEnu = 'N'           
    for i in range(2, 100):                
        BancontactEnu = BancontactEnu + str(i)
        i = i + 1
        Value = sheet1[BancontactEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass   
            elif str(Value) >= "0":
                Bancontact.append(Value)
                MainDump.append(Value)
        BancontactEnu = 'N'

    Cash = []
    CashEnu = 'O'
    for i in range(2, 100):
        CashEnu = CashEnu + str(i)
        i = i + 1
        Value = sheet1[CashEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) != "X":
                Cash.append(Value)
                MainDump.append(Value)
        CashEnu = 'O'

    VoorschotOverschrijving = []
    VOEnu = 'P'
    for i in range(2, 100):
        VOEnu = VOEnu + str(i)
        i = i + 1
        Value = sheet1[VOEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                VoorschotOverschrijving.append(Value)
                MainDump.append(Value)
        VOEnu = 'P'

    Cheque = []
    CEnu = 'Q'
    for i in range(2, 100):
        CEnu = CEnu + str(i)
        i = i + 1
        Value = sheet1[CEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                Cheque.append(Value)
                MainDump.append(Value)
        CEnu = 'Q'


    MainDumpV2 = []

    o = 0
    for i in KlantNrNaam:
        try:
            MainDumpV2.append(str(KlantNrNaam[o]) + "  " + str(Revenue[o]) + "  " + str(Onkosten[o]) + "  " +str(Taks[o]) + "  " +str(AfbetalingLichting[o]) + "  " +str(Provisie[o])+ "  " + str(AfbetalingTaks[o]) + "  " + str(SchuldLichting[o]) + "  " + str(Schuldtaks[o]) + "  " + str(Bancontact[o]) + "  " + str(Cash[o]))
        except:
             pass
        o = o +1
    
    Date = str(sheet1['C2'].value)
    Day = "0"
    if len(Date) == 7:
        Day = "0" + str(Date[0:1])
    else:
        Day = str(Date[0:2])
        
    DateToPrint = Day +"/"+str(Date[-6:-4])+"/"+str(Date[-4:])

    Kasboek = []
    KasboekDebet = []
    KasboekVerwijzing1 = []
    KasboekVerwijzing2 = []
    Kolomreferentie = []
    Klantreferentie = []
    KlantDBNr = []
    NewClientNr = []
    KasboekBetalingLijst = []

    
    wb = Workbook()
    sheet = wb.add_sheet("Kas")

    # geel
    style2 = xlwt.easyxf('font: bold off, color black;\
                     borders: top_color black, bottom_color black, right_color black, left_color black,\
                              left thin, right thin, top thin, bottom thin;\
                     pattern: pattern solid, fore_color yellow;')
    # grijs
    style3 = xlwt.easyxf('font: bold off, color black;\
                     borders: top_color black, bottom_color black, right_color black, left_color black,\
                              left thin, right thin, top thin, bottom thin;\
                     pattern: pattern solid, fore_color gray25;')
    # wit
    style4 = xlwt.easyxf('font: bold off, color black;\
                     borders: top_color black, bottom_color black, right_color black, left_color black,\
                              left thin, right thin, top thin, bottom thin;\
                     pattern: pattern solid, fore_color white;')
    # wit-bold
    style5 = xlwt.easyxf('font: bold on, color black;\
                     borders: top_color black, bottom_color black, right_color black, left_color black,\
                              left no_line, right no_line, top thick, bottom thick;\
                     pattern: pattern solid, fore_color pale_blue;')


    font0 = xlwt.easyfont('')
    font1 = xlwt.easyfont('bold true')
    font2 = xlwt.easyfont('color_index red')
    style = xlwt.easyxf('font: color_index blue')
    style66 =xlwt.easyxf('pattern: pattern solid, fore_colour yellow;')    
    style99 =xlwt.easyxf('pattern: pattern solid, fore_colour light_blue;')

    seg1 = ('Klantnr:', font1)
    seg2 = ('', font2)
    seg3 = ('', font0)
    seg4 = ('', font1)

    eseg1 = ('312381', font2)
    eseg2 = ('', font2)
    eseg3 = ('', font0) 
    eseg4 = ('', font1)
                         




    #Kasdemo
    #Kasregel maken
    NK = str(0)
    KasNrValue = ""
    OndernemingValue = str(sheet1['S1'].value)
    if OndernemingValue == "BL":
        KasNrValue = KN[0]
    elif OndernemingValue == "WMC":
        KasNrValue = KN[1]
    elif OndernemingValue == "NACO":
        KasNrValue = KN[2]
        
    
    NK = str(KasNrValue)
    #NK = NummerKas.get()
    Kasvermelding = "KAS " + str((Date[-2:])) + "-" + str(NK) + "/"

    
    #Bereken Taksen
    TaksV2 = []
    SchuldtaksV2 = []
    
    TaksV2 = Taks.copy()        
    SchuldtaksV2 = Schuldtaks.copy()
        
    del SchuldtaksV2[-1]
    del TaksV2[-1]

    ClientNRTaks = []
    ExpertMNR = ""
    ClientsToReport = ""

    CalculatedTaks = 0
    a = 0
    for i in TaksV2:
        try:
            CalculatedTaks = float(TaksV2[a]) - float(SchuldtaksV2[a])
            if CalculatedTaks == 0:
                a = a+1 
            else:
                KasboekDebet.append("")
                KasboekBetalingLijst.append("Betaling")
                Kasboek.append(float(CalculatedTaks))
                KasboekVerwijzing1.append("CFBING")
                KasboekVerwijzing2.append("TAKSEN")
                Kolomreferentie.append("Klant")
                Klantreferentie.append(str(KlantNrNaam[a]))
                KlantDBNr.append(str(KlantNrNaam[a][0:6]))
                u = 0
                for t in ExpertMDB:
                    if str(KlantNrNaam[a][0:6]) == str(ExpertMDB[u]):
                        ExpertMNR = ExpertMDB[u]                       
                        ClientNRTaks.append(str(KlantDB[u]))
                    u = u +1
                                            
                if len(ClientNRTaks)>1:
                    ClientsToReport = ClientsToReport + str(ExpertMNR) + "\n"  
                                            
                NewClientNr.append(ClientNRTaks[-1])

                ClientNRTaks.clear()
  
                        
                a = a+1
        except:
            pass

    if len(ClientsToReport) > 0:
        messagebox.showinfo("1+ klant", ClientsToReport)   

    #Bereken Lichting
    AfbetalingLichtingV2 = []
    AfbetalingLichtingV2 = AfbetalingLichting.copy()
    del AfbetalingLichtingV2[-1]
    
    r = 0
    for i in AfbetalingLichtingV2:
        if AfbetalingLichtingV2[r] == 0:
            r = r +1
        else:
            KasboekDebet.append("")
            Kasboek.append(float(AfbetalingLichtingV2[r]))
            KasboekVerwijzing1.append("")
            KasboekVerwijzing2.append("")
            KasboekBetalingLijst.append("Betaling")
            Kolomreferentie.append("Klant")
            Klantreferentie.append(str(KlantNrNaam[r]))
            KlantDBNr.append(str(KlantNrNaam[r][0:6]))
            u = 0
            for t in ExpertMDB:
                if str(KlantNrNaam[r][0:6]) == str(ExpertMDB[u]):
                    NewClientNr.append(str(KlantDB[u]))
                u = u +1
            r = r+1

    #Berekenen Provisie
    ProvisieV2 = []
    ProvisieV2 = Provisie.copy()
    del ProvisieV2[-1]

    r = 0
    for i in ProvisieV2:
        if ProvisieV2[r] == 0:
            r = r +1
        else:
            KasboekDebet.append("")
            Kasboek.append(float(ProvisieV2[r]))
            KasboekVerwijzing1.append("")
            KasboekVerwijzing2.append("PROVISIES")
            KasboekBetalingLijst.append("Betaling")
            Kolomreferentie.append("Klant")
            Klantreferentie.append(str(KlantNrNaam[r]))
            KlantDBNr.append(str(KlantNrNaam[r][0:6]))
            u = 0
            for t in ExpertMDB:
                if str(KlantNrNaam[r][0:6]) == str(ExpertMDB[u]):
                    NewClientNr.append(str(KlantDB[u]))
                u = u +1
            r = r+1

    #Berekening Afbetaling Taks
    AfbetalingTaksV2 = []
    AfbetalingTaksV2 = AfbetalingTaks.copy()
    del AfbetalingTaksV2[-1]

    r = 0
    for i in AfbetalingTaksV2:
        if AfbetalingTaksV2[r] == 0:
            r = r +1
        else:
            KasboekDebet.append("")
            Kasboek.append(float(AfbetalingTaksV2[r]))
            KasboekVerwijzing1.append("CFBING")
            KasboekVerwijzing2.append("TAKSEN")
            KasboekBetalingLijst.append("Betaling")
            Kolomreferentie.append("Klant")
            Klantreferentie.append(str(KlantNrNaam[r]))
            KlantDBNr.append(str(KlantNrNaam[r][0:6]))
            u = 0
            for t in ExpertMDB:
                if str(KlantNrNaam[r][0:6]) == str(ExpertMDB[u]):
                    NewClientNr.append(str(KlantDB[u]))
                u = u +1
            r = r+1

    #Berekenen Schuld Lichting
    SchuldLichtingV2 = []
    SchuldLichtingV2 = SchuldLichting.copy()
    del SchuldLichtingV2[-1]
    #foutoplossing Schuld Lichting tov Cash
    y = 0
    for yy in SchuldLichtingV2:
        if SchuldLichtingV2[y] == 0:
            pass
            y = y +1
        elif SchuldLichtingV2[y] < Cash[y]:
            KasboekDebet.append(float(SchuldLichtingV2[y]))                   
            KasboekVerwijzing1.append("")
            KasboekVerwijzing2.append("")
            KasboekBetalingLijst.append("")
            Kolomreferentie.append("Klant")
            Klantreferentie.append(str(KlantNrNaam[y] + " - factuur"))
            KlantDBNr.append("")
            u = 0
            for t in ExpertMDB:
                if str(KlantNrNaam[y][0:6]) == str(ExpertMDB[u]):
                    NewClientNr.append(str(KlantDB[u]))
                u = u +1
            Kasboek.append("")

            KasboekDebet.append(float(Onkosten[y]))                   
            KasboekVerwijzing1.append("")
            KasboekVerwijzing2.append("")
            KasboekBetalingLijst.append("")
            Kolomreferentie.append("Klant")
            Klantreferentie.append(str(KlantNrNaam[y] + " - onkosten"))
            KlantDBNr.append("")
            u = 0
            for t in ExpertMDB:
                if str(KlantNrNaam[y][0:6]) == str(ExpertMDB[u]):
                    NewClientNr.append(str(KlantDB[u]))
                u = u +1
            Kasboek.append("")

            
            SchuldLichtingV2[y] = 0            
            y = y +1     
    
    r = 0
    for i in SchuldLichtingV2:
        if SchuldLichtingV2[r] == 0:
            pass
            r = r+1
        elif Revenue[r] - SchuldLichtingV2[r] - Onkosten[r] == 0:
            pass
            r = r+1
        elif Revenue[r] - SchuldLichtingV2[r] - Onkosten[r] > 0 and Revenue[r] - SchuldLichtingV2[r] - Onkosten[r] < 0.10:
            pass
            r = r+1
        elif Revenue[r] - SchuldLichtingV2[r] - Onkosten[r] < 0 and Revenue[r] - SchuldLichtingV2[r] - Onkosten[r] > -0.10:
            pass
            r = r+1
        else:
            if float(Revenue[r] - SchuldLichtingV2[r] - Onkosten[r]) > 0: 
                KasboekDebet.append("")
                Kasboek.append(float(Revenue[r] - SchuldLichtingV2[r] - Onkosten[r]))
            else:
                KasboekDebet.append(float(Revenue[r] - SchuldLichtingV2[r] - Onkosten[r])*-1)
                Kasboek.append("")                
            KasboekVerwijzing1.append("")
            KasboekVerwijzing2.append("")
            KasboekBetalingLijst.append("Betaling")
            Kolomreferentie.append("Klant")
            Klantreferentie.append(str(KlantNrNaam[r]))
            KlantDBNr.append(str(KlantNrNaam[r][0:6]))
            u = 0
            for t in ExpertMDB:
                if str(KlantNrNaam[r][0:6]) == str(ExpertMDB[u]):
                    NewClientNr.append(str(KlantDB[u]))
                u = u +1
            r = r+1

    #Invullen Voorschot overschrijving
    VoorschotOverschrijvingV2 = []
    VoorschotOverschrijvingV2 = VoorschotOverschrijving.copy()
    del VoorschotOverschrijvingV2[-1]
    
    r = 0
    for i in VoorschotOverschrijvingV2:
        if VoorschotOverschrijvingV2[r] == 0:
            r = r +1
        else:
            KasboekDebet.append(float(VoorschotOverschrijvingV2[r]))
            Kasboek.append("")
            KasboekVerwijzing1.append("")
            KasboekVerwijzing2.append("")
            KasboekBetalingLijst.append("Betaling")
            Kolomreferentie.append("Klant")
            Klantreferentie.append(str(KlantNrNaam[r]))
            KlantDBNr.append(str(KlantNrNaam[r][0:6]))
            u = 0
            for t in ExpertMDB:
                if str(KlantNrNaam[r][0:6]) == str(ExpertMDB[u]):
                    NewClientNr.append(str(KlantDB[u]))
                u = u +1
            r = r+1

    #KasCash invullen
    KasCshValue = 0.0
    KasCshValu = Saldo570110.get()
    KasCshValue = KasCshValu.replace(",",".")
    
    try:
        Kasboek.append(float(KasCshValue))
    except:
        messagebox.showinfo("No Go","KasCash is empty!")
        sys.exit(1)
        
    KasboekDebet.append("")
    KasboekVerwijzing1.append("")
    KasboekVerwijzing2.append("")
    KasboekBetalingLijst.append("")
    Kolomreferentie.append("Bank")
    Klantreferentie.append(DateToPrint)
    KlantDBNr.append("")
    NewClientNr.append("KASCASH")

    #Berekenen input bancontact
    BancontactV2 = []
    BancontactV2 = Bancontact.copy()
    del BancontactV2[-1]
    j = 0
    for i in BancontactV2:
        if BancontactV2[j] == 0:
            j = j+1

        else:
            KasboekDebet.append(float(BancontactV2[j]))
            j = j+1                    
            KasboekVerwijzing1.append("")
            KasboekVerwijzing2.append("")
            KasboekBetalingLijst.append("")
            Kolomreferentie.append("Grootboekrekening")
            Klantreferentie.append(DateToPrint)
            KlantDBNr.append("")
            NewClientNr.append("580200")
            Kasboek.append("")
            
    #Bekomen Bedragen Biljetten
    Biljetten = [] 
    B0 = Biljetten00.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten01.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten02.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten03.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten04.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten05.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten06.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten07.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten08.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten09.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten10.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten11.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten12.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten13.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten14.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten15.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten16.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten17.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten18.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten19.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten20.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten21.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten22.get()
    if B0 != "":
        Biljetten.append(B0)
    B0 = Biljetten23.get()
    if B0 != "":
        Biljetten.append(B0)

    b = 0
    for i in Biljetten:
        KasboekDebet.append(float(Biljetten[b]))
        b = b+1                    
        KasboekVerwijzing1.append("")
        KasboekVerwijzing2.append("")
        KasboekBetalingLijst.append("")
        Kolomreferentie.append("Grootboekrekening")
        Klantreferentie.append(DateToPrint)
        KlantDBNr.append("")
        NewClientNr.append("580000")
        Kasboek.append("")

    #Bekomen Bedragen Munten
    Munten = [] 
    M00 = Munten0.get()
    M0 = M00.replace(",",".")    
    if M0 != "":
        Munten.append(M0)
    M00 = Munten1.get()
    M0 = M00.replace(",",".") 
    if M0 != "":
        Munten.append(M0)
    M00 = Munten2.get()
    M0 = M00.replace(",",".") 
    if M0 != "":
        Munten.append(M0)
        
    m = 0
    for i in Munten:
        KasboekDebet.append(float(Munten[m]))
        m = m+1                    
        KasboekVerwijzing1.append("")
        KasboekVerwijzing2.append("")
        KasboekBetalingLijst.append("")
        Kolomreferentie.append("Grootboekrekening")
        Klantreferentie.append(DateToPrint)
        KlantDBNr.append("")
        NewClientNr.append("580210")
        Kasboek.append("")

    #Speciale ingave opnemen
    SpecDebet = []
    try:
        SID00 = SID0.get()
        SID000 = SID00.replace(",",".")
        SpecDebet.append(float(SID000))
    except:
        pass
    try:
        SID01 = SID1.get()
        SID001 = SID01.replace(",",".")
        SpecDebet.append(float(SID001))
    except:
        pass
    try:
        SID02 = SID2.get()
        SID002 = SID02.replace(",",".")
        SpecDebet.append(float(SID002))
    except:
        pass
    
    SpecCredit = []
    try:
        SIC00 = SIC0.get()
        SIC000 = SIC00.replace(",",".")
        SpecCredit.append(float(SIC000))
    except:
        pass
    try:
        SIC01 = SIC1.get()
        SIC001 = SIC01.replace(",",".")
        SpecCredit.append(float(SIC001))
    except:
        pass
    try:
        SIC02 = SIC2.get()
        SIC002 = SIC02.replace(",",".")
        SpecCredit.append(float(SIC002))
    except:
        pass

    m = 0
    for u in SpecDebet:
        KasboekDebet.append(float(SpecDebet[m]))                
        KasboekVerwijzing1.append("")
        KasboekVerwijzing2.append("")
        KasboekBetalingLijst.append("")
        Kolomreferentie.append("Grootboekrekening")
        Klantreferentie.append(DateToPrint)
        KlantDBNr.append("Commentaar voorzien")
        NewClientNr.append("580000")
        Kasboek.append("")
        m = m+1   

    m = 0
    for u in SpecCredit:
        KasboekDebet.append("")               
        KasboekVerwijzing1.append("")
        KasboekVerwijzing2.append("")
        KasboekBetalingLijst.append("")
        Kolomreferentie.append("Grootboekrekening")
        Klantreferentie.append(DateToPrint)
        KlantDBNr.append("Commentaar voorzien")
        NewClientNr.append("580000")
        Kasboek.append(float(SpecCredit[m]))
        m = m+1     


    #Betalingsverschil invullen
    p = 0
    SumKasboek = 0.0
    for i in Kasboek:
        try:
            SumKasboek = SumKasboek + float(Kasboek[p])
            p = p + 1
        except:
            p = p+1

    AnswerCredit.grid(row=7,column=7)
    AnswerCredit.config(text =(str("%.2f" % SumKasboek)))
                       
    p = 0
    
    SumKasboekDebet = 0.0
    for i in KasboekDebet:
        try:
            SumKasboekDebet = SumKasboekDebet + float(KasboekDebet[p])
            p = p + 1
        except:
            p = p+1

    AnswerDebet.grid(row=7,column=6)
    AnswerDebet.config(text =(str("%.2f" % SumKasboekDebet)))

    BDF = SumKasboekDebet - SumKasboek
    if BDF >= 0:        
        Kasboek.append(float(BDF))
        KasboekDebet.append("")
        NewClientNr.append("758400")
        if float(BDF) > 0.15:
            messagebox.showinfo("Value notice","Betalingsverschil is mogelijk te groot")
            
    else:
        Kasboek.append("")
        KasboekDebet.append(float(BDF)*-1)
        NewClientNr.append("658400")
        if float(BDF)*-1 > 0.15:
            messagebox.showinfo("Value notice","Betalingsverschil is mogelijk te groot")
        

    AnswerKasV.grid(row=9,column=6)
    AnswerKasV.config(text =(str("%.2f" % BDF)))

        
    KasboekVerwijzing1.append("SHARFI")
    KasboekVerwijzing2.append("")
    KasboekBetalingLijst.append("")
    Kolomreferentie.append("Grootboekrekening")
    Klantreferentie.append(DateToPrint)
    KlantDBNr.append("")

    

  
        
    #Kasboek schrijven
    leadingnumberforzero = 1
    m = 0
    for i in Kolomreferentie:
        sheet.write(m,0, DateToPrint)
        sheet.write(m,1, DateToPrint)
        sheet.write(m,8, Kasboek[m])
        sheet.write(m,9, str(KasboekVerwijzing1[m]))
        sheet.write(m,10, str(KasboekVerwijzing2[m]))
        sheet.write(m,4, str(Kolomreferentie[m]))
        strlnfz = str(leadingnumberforzero)
        fillednumber = strlnfz.zfill(3)
        sheet.write(m,3, str(Kasvermelding)+fillednumber)
        leadingnumberforzero = leadingnumberforzero +1
        sheet.write(m,6, str(Klantreferentie[m]))
        sheet.write(m,7, KasboekDebet[m])
        sheet.write(m,11, str(KlantDBNr[m]))
        sheet.write(m,5, str(NewClientNr[m]))
        sheet.write(m,2, str(KasboekBetalingLijst[m]))
        m = m+1
    first_col0 = sheet.col(6)
    first_col0.width = 256 * 40
    first_col1 = sheet.col(3)
    first_col1.width = 256 * 20
    first_col2 = sheet.col(0)
    first_col2.width = 256 * 11
    first_col3 = sheet.col(1)
    first_col3.width = 256 * 11
    first_col4 = sheet.col(12)
    first_col4.width = 256 * 11
    first_col5 = sheet.col(5)
    first_col5.width = 256 * 11

    
    #Balk met sommen
    p = 0
    SumKasboek = 0.0
    for i in Kasboek:
        try:
            SumKasboek = SumKasboek + float(Kasboek[p])
            p = p + 1
        except:
            p = p+1
    p = 0
    SumKasboekDebet = 0.0
    for i in Kasboek:
        try:
            SumKasboekDebet = SumKasboekDebet + float(KasboekDebet[p])
            p = p + 1
        except:
            p = p+1

    DifDebetCredit = SumKasboek - SumKasboekDebet
        
        
    SumBar = [" "," "," "," "," "," "," ","Eindsaldo",str(DifDebetCredit),"DIFF",0," "]
    
    EndNr = len(Kasboek)
    o = 0
    for i in SumBar:
        sheet.write_rich_text(int(EndNr),o, (str(i), " ", " ", ''), style5)
        o = o+1
    #Datum in het geel ter referentie
    sheet.write_rich_text(0, 12, (DateToPrint, seg2, seg3, ''), style66)

    #Prepare dump
    CountRowsO = 0
    SA = []
    StringShowAll = ""
    for x in MainDumpV2:
        CountRowsO = CountRowsO + 1
        SA.append(str(x))
        StringShowAll = StringShowAll + str(x) + "\n"
 
    for i in range(0,CountRowsO):
        listbox.insert(END, str(SA[i]))
    listbox.pack(side=LEFT, fill=BOTH, ipadx = 660)
    
    try:
        wb.save('KasbladDemo.xls')
        file = 'KasbladDemo.xls'
        os.startfile(file)
    except:
        messagebox.showinfo("No Go","Verwijder KasbladDemo uit je map")
        sys.exit(1)

        
    #Delete all lists
    MainDump.clear()
    KlantNrNaam.clear()
    Taks.clear()
    Schuldtaks.clear()
    Provisie.clear()
    AfbetalingTaks.clear()
    SchuldLichting.clear()
    AfbetalingLichting.clear()
    Revenue.clear()
    Onkosten.clear()
    Bancontact.clear()
    Cash.clear()
    VoorschotOverschrijving.clear()
    Cheque.clear()
    MainDumpV2.clear()
    Kasboek.clear()
    KasboekDebet.clear()
    KasboekVerwijzing1.clear()
    KasboekVerwijzing2.clear()
    Kolomreferentie.clear()
    Klantreferentie.clear()
    KlantDBNr.clear()
    NewClientNr.clear()
    KasboekBetalingLijst.clear()
    TaksV2.clear()
    SchuldtaksV2.clear()
    AfbetalingLichtingV2.clear()
    ProvisieV2.clear()
    AfbetalingTaksV2.clear()
    SchuldLichtingV2.clear()
    VoorschotOverschrijvingV2.clear()
    Biljetten.clear()
    Munten.clear()
    

def CheckEW():
    wb1 = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\Kasblad.xlsx')
    sheet1 = wb1['Sheet1']

    Revenue = []
    RevenueEnu = 'F'
    for i in range(2, 100):
        RevenueEnu = RevenueEnu + str(i)
        i = i + 1
        Value = sheet1[RevenueEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                Revenue.append(Value)
                MainDump.append(Value)
        RevenueEnu = 'F'

    Onkosten = []
    OnkostenEnu = 'G'
    for i in range(2, 100):
        OnkostenEnu = OnkostenEnu + str(i)
        i = i + 1
        Value = sheet1[OnkostenEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) >= "0":
                Onkosten.append(Value)
                MainDump.append(Value)
        OnkostenEnu = 'G'
        
    WinningsInpu = Saldo7004.get()
    WinningsInput = WinningsInpu.replace(",",".")
    ExpensensInpu = Saldo613822.get()
    ExpensensInput = ExpensensInpu.replace(",",".")
    WinningsImport = Revenue[-1]
    ExpensesImport = Onkosten[-1]

    AnswerLabel = Label(page1, text="...")
    AnswerLabel.grid(row=6,column=1)
    AnswerLabel.config(text =(str("Calculating")))

    

    try:
        WinningsDifference = float(WinningsInput) - float(WinningsImport)
        ExpensesDifference = float(ExpensensInput) - float(ExpensesImport)
        AnswerLabel.grid(row=6,column=1)
        AnswerLabel.config(text =(str("Winnings Difference = " + str("%.2f" % WinningsDifference) + "\n" + "Expenses Difference = " + str("%.2f" % ExpensesDifference))))         
    except:
        AnswerLabel.grid(row=6,column=1)
        AnswerLabel.config(text =(str("Geen input!")))

def ExcelOK():
    wb99 = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\Teller.xlsx', read_only=True)
    sheet99 = wb99['Data']

    KN = []
    KNenu = 'B'
    k = 2
    for i in range(2,5):
        try:
            KNenu = KNenu+str(k)
            KN.append(str(sheet99[KNenu].value))
            KNenu = 'B'
            k = k + 1
        except:
            pass

    wb1 = load_workbook(filename = r'C:\Users\...\OneDrive - Gamesgroup\Documenten\Kasblad.xlsx')
    sheet1 = wb1['Sheet1']
    OndernemingValue = str(sheet1['S1'].value)

    if OndernemingValue == "BL":
        KasNrValue = KN[0]
    elif OndernemingValue == "WMC":
        KasNrValue = KN[1]
    elif OndernemingValue == "NACO":
        KasNrValue = KN[2]
        
    AnswerLabel.grid(row=6,column=1)
    AnswerLabel.config(text =(""))
    AnswerCredit.grid(row=7,column=7)
    AnswerCredit.config(text =(""))
    AnswerDebet.grid(row=7,column=6)
    AnswerDebet.config(text =(""))
    AnswerKasV.grid(row=9,column=6)
    AnswerKasV.config(text =(""))


    xfile = openpyxl.load_workbook(r'C:\Users\...\OneDrive - Gamesgroup\Documenten\Teller.xlsx')
    sheet =  xfile.get_sheet_by_name('Data')

    if OndernemingValue == "BL":
        sheet['B2'] = int(KasNrValue)+1
    elif OndernemingValue == "WMC":
        sheet['B3'] = int(KasNrValue)+1
    elif OndernemingValue == "NACO":
        sheet['B4'] = int(KasNrValue)+1
        
    xfile.save(r'C:\Users\p.pauwels\OneDrive - Gamesgroup\Documenten\Teller.xlsx')

def Csheet():
    wb1 = load_workbook(filename = r'C:\Users\p.pauwels\OneDrive - Gamesgroup\Documenten\Kasblad.xlsx')
    sheet1 = wb1['Sheet1']
    
    ListOne = []
    ListTwo = []
    ListThree = []
    ListFour = []
    ListFive = []
    ListSix = []
    ListSeven = []
    ListEight = []
    ListNine = []
    ListTen = []

    BiljettenSUM = [] 
    B0 = Biljetten00.get()
    B0SUM = Biljetten00SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten01.get()
    B0SUM = Biljetten01SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten02.get()
    B0SUM = Biljetten02SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten03.get()
    B0SUM = Biljetten03SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten04.get()
    B0SUM = Biljetten04SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten05.get()
    B0SUM = Biljetten05SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten06.get()
    B0SUM = Biljetten06SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten07.get()
    B0SUM = Biljetten07SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten08.get()
    B0SUM = Biljetten08SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten09.get()
    B0SUM = Biljetten09SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten10.get()
    B0SUM = Biljetten10SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten11.get()
    B0SUM = Biljetten11SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten12.get()
    B0SUM = Biljetten12SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten13.get()
    B0SUM = Biljetten13SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten14.get()
    B0SUM = Biljetten14SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten15.get()
    B0SUM = Biljetten15SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten16.get()
    B0SUM = Biljetten16SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten17.get()
    B0SUM = Biljetten17SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten18.get()
    B0SUM = Biljetten18SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten19.get()
    B0SUM = Biljetten19SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten20.get()
    B0SUM = Biljetten20SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten21.get()
    B0SUM = Biljetten21SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten22.get()
    B0SUM = Biljetten22SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
    B0 = Biljetten23.get()
    B0SUM = Biljetten23SUM.get()
    if B0 != "":
        BiljettenSUM.append([B0,B0SUM])
        
    MuntenSUM = [] 
    M00 = Munten0.get()
    M00SUM = Munten0SUM.get()
    M0SUM = M00SUM.replace(",",".")
    M0 = M00.replace(",",".")    
    if M0 != "":
        MuntenSUM.append([M0,M0SUM])
    M00 = Munten1.get()
    M00SUM = Munten1SUM.get()
    M0SUM = M00SUM.replace(",",".")
    M0 = M00.replace(",",".") 
    if M0 != "":
       MuntenSUM.append([M0,M0SUM])
    M00 = Munten2.get()
    M00SUM = Munten2SUM.get()
    M0SUM = M00SUM.replace(",",".")
    M0 = M00.replace(",",".") 
    if M0 != "":
       MuntenSUM.append([M0,M0SUM])


    for i in MuntenSUM: 
        if str(i[1]) == "1": 
            ListOne.append(i[0]) 
        elif str(i[1]) == "2":
            ListTwo.append(i[0])
        elif str(i[1]) == "3": 
            ListThree.append(i[0]) 
        elif str(i[1]) == "4":
            ListFour.append(i[0])
        elif str(i[1]) == "5":
            ListFive.append(i[0])
        elif str(i[1]) == "6": 
            ListSix.append(i[0]) 
        elif str(i[1]) == "7":
            ListSeven.append(i[0])
        elif str(i[1]) == "8": 
            ListEight.append(i[0]) 
        elif str(i[1]) == "9":
            ListNine.append(i[0])
        elif str(i[1]) == "10":
            ListTen.append(i[0])

    for i in BiljettenSUM: 
        if str(i[1]) == "1": 
            ListOne.append(i[0]) 
        elif str(i[1]) == "2":
            ListTwo.append(i[0])
        elif str(i[1]) == "3": 
            ListThree.append(i[0]) 
        elif str(i[1]) == "4":
            ListFour.append(i[0])
        elif str(i[1]) == "5":
            ListFive.append(i[0])
        elif str(i[1]) == "6": 
            ListSix.append(i[0]) 
        elif str(i[1]) == "7":
            ListSeven.append(i[0])
        elif str(i[1]) == "8": 
            ListEight.append(i[0]) 
        elif str(i[1]) == "9":
            ListNine.append(i[0])
        elif str(i[1]) == "10":
            ListTen.append(i[0])

    SumLists = []
    Sum1 = 0
    for m in ListOne:
        Sum1 = Sum1 + int(m)
    SumLists.append(Sum1)
    Sum2 = 0
    for m in ListTwo:
        Sum2 = Sum2 + int(m)
    SumLists.append(Sum2)
    Sum3 = 0
    for m in ListThree:
        Sum3 = Sum3 + int(m)
    SumLists.append(Sum3)
    Sum4 = 0
    for m in ListFour:
        Sum4 = Sum4 + int(m)
    SumLists.append(Sum4)
    Sum5 = 0
    for m in ListFive:
        Sum5 = Sum5 + int(m)
    SumLists.append(Sum5)
    Sum6 = 0
    for m in ListSix:
        Sum6 = Sum6 + int(m)
    SumLists.append(Sum6)
    Sum7 = 0
    for m in ListSeven:
        Sum7 = Sum7 + int(m)
    SumLists.append(Sum7)
    Sum8 = 0
    for m in ListEight:
        Sum8 = Sum8 + int(m)
    SumLists.append(Sum8)
    Sum9 = 0
    for m in ListNine:
        Sum9 = Sum9 + int(m)
    SumLists.append(Sum9)
    Sum10 = 0
    for m in ListTen:
        Sum10 = Sum10 + int(m)
    SumLists.append(Sum10)

    print(SumLists)
    
    



    Bancontact = []
    BancontactEnu = 'N'           
    for i in range(2, 100):                
        BancontactEnu = BancontactEnu + str(i)
        i = i + 1
        Value = sheet1[BancontactEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass   
            elif str(Value) >= "0":
                Bancontact.append(Value)
                MainDump.append(Value)
        BancontactEnu = 'N'

    Cash = []
    CashEnu = 'O'
    for i in range(2, 100):
        CashEnu = CashEnu + str(i)
        i = i + 1
        Value = sheet1[CashEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) != "X":
                Cash.append(Value)
                MainDump.append(Value)
        CashEnu = 'O'

    Lichter = []
    UniekeLichters = []
    UL = []
    LichterEnu = 'D'
    for i in range(2, 100):
        LichterEnu = LichterEnu + str(i)
        i = i + 1
        Value = sheet1[LichterEnu].value
        try:
            CheckVal = Value[0:3]
        except:
            CheckVal = "OK"
        if Value is not None:
            if CheckVal == "=SU":
                pass
            elif str(Value) != "X":
                Lichter.append(Value)
                MainDump.append(Value)
        LichterEnu = 'D'

    for u in Lichter:
        if u in UL:
            pass
        else:
            UniekeLichters.append([u,0])
            UL.append(u)

    for pp in UniekeLichters:
        Enu = 0
        for o in Lichter:
            if o == pp[0]:
                CalcValue = int(pp[1])
                print(str(pp[1]))
                CalcValue = CalcValue + int(Cash[Enu])
                pp[1] = CalcValue
            Enu = Enu + 1         
                
 
    
"""Opmaak Tabs"""
     
# gives weight to the cells in the grid
rows = 0
while rows < 50:
    main.rowconfigure(rows, weight=1)
    main.columnconfigure(rows, weight=1)
    rows += 1
 
# Defines and places the notebook widget
nb = ttk.Notebook(main)
nb.grid(row=1, column=0, columnspan=50, rowspan=49, sticky='NESW')
 
# Adds tab 1 of the notebook
page1 = ttk.Frame(nb)
nb.add(page1, text='KasMaker')

"""Opmaak 1ste tab"""
#Setting it up
img = ImageTk.PhotoImage(Image.open("caisse.png"))
imglabel1 = Label(page1, image=img, wraplength=5).grid(row=1, column=8)


BT0 = tkinter.Button(page1, text="Create Excel", command=CreateExcel).grid(row=12,column=6)
BT1 = tkinter.Button(page1, text="Check expenses/winnings", command=CheckEW).grid(row=9,column=1)
BT2 = tkinter.Button(page1, text="Excel OK", command=ExcelOK).grid(row=14,column=6)
BT3 = tkinter.Button(page1, text="Controlesheet", command=Csheet).grid(row=16,column=6)
       
#Displaying it
tkinter.Label(page1, text="Saldo 7004 ",borderwidth=1 ).grid(row=1,column=1)
Saldo7004 = tkinter.Entry(page1,borderwidth=1, width =20 )
Saldo7004.grid(row=2,column=1)
tkinter.Label(page1, text=" ",borderwidth=1 ).grid(row=3,column=1)
tkinter.Label(page1, text="Saldo 613822 ",borderwidth=1 ).grid(row=4,column=1)
Saldo613822 = tkinter.Entry(page1,borderwidth=1, width =20 )
Saldo613822.grid(row=5,column=1)

AnswerLabel = Label(page1, text="...")
AnswerLabel.grid(row=6,column=1)

tkinter.Label(page1, text="Saldo 570110 ",borderwidth=1 ).grid(row=7,column=1)
Saldo570110 = tkinter.Entry(page1,borderwidth=1, width =20 )
Saldo570110.grid(row=8,column=1)

#tkinter.Label(page1, text="Nummering",borderwidth=1 ).grid(row=10,column=1)
#NummerKas = tkinter.Entry(page1,borderwidth=1, width =20 )
#NummerKas.grid(row=11,column=1)

#tkinter.Label(page1, text="Bestandsnaam",borderwidth=1 ).grid(row=12,column=1)
#Bestandsnaam = tkinter.Entry(page1,borderwidth=1, width =20 )
#Bestandsnaam.grid(row=13,column=1)
#Bestandsnaam.insert(0,"Kasblad")

tkinter.Label(page1, text="          ",borderwidth=1 ).grid(row=1,column=2)

tkinter.Label(page1, text="Ingave munten",borderwidth=1 ).grid(row=1,column=3)
Munten0 = tkinter.Entry(page1,borderwidth=1, width =20 )
Munten0.grid(row=2,column=3)
Munten1 = tkinter.Entry(page1,borderwidth=1, width =20 )
Munten1.grid(row=3,column=3)
Munten2 = tkinter.Entry(page1,borderwidth=1, width =20 )
Munten2.grid(row=4,column=3)
tkinter.Label(page1, text="          ",borderwidth=1 ).grid(row=5,column=3)
tkinter.Label(page1, text="Ingave biljetten",borderwidth=1 ).grid(row=6,column=3)
Biljetten00 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten00.grid(row=7,column=3)
Biljetten01 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten01.grid(row=8,column=3)
Biljetten02 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten02.grid(row=9,column=3)
Biljetten03 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten03.grid(row=10,column=3)
Biljetten04 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten04.grid(row=11,column=3)
Biljetten05 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten05.grid(row=12,column=3)
Biljetten06 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten06.grid(row=13,column=3)
Biljetten07 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten07.grid(row=14,column=3)
Biljetten08 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten08.grid(row=15,column=3)
Biljetten09 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten09.grid(row=16,column=3)
Biljetten10 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten10.grid(row=17,column=3)
Biljetten11 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten11.grid(row=18,column=3)
Biljetten12 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten12.grid(row=19,column=3)
Biljetten13 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten13.grid(row=20,column=3)
Biljetten14 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten14.grid(row=21,column=3)
Biljetten15 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten15.grid(row=22,column=3)
Biljetten16 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten16.grid(row=23,column=3)
Biljetten17 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten17.grid(row=24,column=3)
Biljetten18 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten18.grid(row=25,column=3)
Biljetten19 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten19.grid(row=26,column=3)
Biljetten20 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten20.grid(row=27,column=3)
Biljetten21 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten21.grid(row=28,column=3)
Biljetten22 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten22.grid(row=29,column=3)
Biljetten23 = tkinter.Entry(page1,borderwidth=1, width =20 )
Biljetten23.grid(row=30,column=3)

Munten0SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Munten0SUM.grid(row=2,column=4)
Munten1SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Munten1SUM.grid(row=3,column=4)
Munten2SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Munten2SUM.grid(row=4,column=4)
Biljetten00SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten00SUM.grid(row=7,column=4)
Biljetten01SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten01SUM.grid(row=8,column=4)
Biljetten02SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten02SUM.grid(row=9,column=4)
Biljetten03SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten03SUM.grid(row=10,column=4)
Biljetten04SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten04SUM.grid(row=11,column=4)
Biljetten05SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten05SUM.grid(row=12,column=4)
Biljetten06SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten06SUM.grid(row=13,column=4)
Biljetten07SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten07SUM.grid(row=14,column=4)
Biljetten08SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten08SUM.grid(row=15,column=4)
Biljetten09SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten09SUM.grid(row=16,column=4)
Biljetten10SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten10SUM.grid(row=17,column=4)
Biljetten11SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten11SUM.grid(row=18,column=4)
Biljetten12SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten12SUM.grid(row=19,column=4)
Biljetten13SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten13SUM.grid(row=20,column=4)
Biljetten14SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten14SUM.grid(row=21,column=4)
Biljetten15SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten15SUM.grid(row=22,column=4)
Biljetten16SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten16SUM.grid(row=23,column=4)
Biljetten17SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten17SUM.grid(row=24,column=4)
Biljetten18SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten18SUM.grid(row=25,column=4)
Biljetten19SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten19SUM.grid(row=26,column=4)
Biljetten20SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten20SUM.grid(row=27,column=4)
Biljetten21SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten21SUM.grid(row=28,column=4)
Biljetten22SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten22SUM.grid(row=29,column=4)
Biljetten23SUM = tkinter.Entry(page1,borderwidth=1, width =5 )
Biljetten23SUM.grid(row=30,column=4)

tkinter.Label(page1, text="     ",borderwidth=1 ).grid(row=1,column=5)

tkinter.Label(page1, text="Speciale ingave",borderwidth=1 ).grid(row=1,column=6)
tkinter.Label(page1, text="Debet",borderwidth=1 ).grid(row=2,column=6)
SID0 = tkinter.Entry(page1,borderwidth=1, width =20 )
SID0.grid(row=3,column=6)
SID1 = tkinter.Entry(page1,borderwidth=1, width =20 )
SID1.grid(row=4,column=6)
SID2 = tkinter.Entry(page1,borderwidth=1, width =20 )
SID2.grid(row=5,column=6)

tkinter.Label(page1, text="Credit",borderwidth=1 ).grid(row=2,column=7)
SIC0 = tkinter.Entry(page1,borderwidth=1, width =20 )
SIC0.grid(row=3,column=7)
SIC1 = tkinter.Entry(page1,borderwidth=1, width =20 )
SIC1.grid(row=4,column=7)
SIC2 = tkinter.Entry(page1,borderwidth=1, width =20 )
SIC2.grid(row=5,column=7)

tkinter.Label(page1, text="Debet na calculatie",borderwidth=1 ).grid(row=6,column=6)
AnswerDebet = Label(page1, text="...")
AnswerDebet.grid(row=7,column=6)
tkinter.Label(page1, text="Credit na calculatie",borderwidth=1 ).grid(row=6,column=7)
AnswerCredit = Label(page1, text="...")
AnswerCredit.grid(row=7,column=7)

tkinter.Label(page1, text="KasVerschil",borderwidth=1 ).grid(row=8,column=6)
AnswerKasV = Label(page1, text="...")
AnswerKasV.grid(row=9,column=6)

# Adds tab 2 of the notebook
page2 = ttk.Frame(nb)
nb.add(page2, text='Dump')



scrollbar = Scrollbar(page2)
scrollbar.pack(side=RIGHT, fill=Y)

listbox = Listbox(page2, yscrollcommand=scrollbar.set)


listbox.pack(side=LEFT, fill=BOTH, ipadx = 660)

scrollbar.config(command=listbox.yview)

 
main.mainloop()
